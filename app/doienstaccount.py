from google.oauth2 import service_account
from googleapiclient.discovery import build
from datetime import datetime, time, timedelta
from dateutil import parser
from zoneinfo import ZoneInfo
import calendar
from Excel_eintrag import excel_setter, excel_setter_homeoffice_p
import json
from openpyxl import load_workbook
from pathlib import Path
from dotenv import load_dotenv
import os
import time
from Invoice_creater import FastBill_invoice_creator
from path_resolver import PathResolver  # Neu: Import der PathResolver-Klasse

load_dotenv()
vor_pro = dict()
# Neu: PathResolver initialisieren
PROJECT_ROOT = Path(__file__).resolve().parent.parent
path_resolver = PathResolver(PROJECT_ROOT)

with open(path_resolver.config_path) as f:
    config = json.load(f)

FastBill_invoice_creator_instance = FastBill_invoice_creator(project_root=path_resolver.fastbill_config)


CALENDAR_ID = config["calendar_id"]
TAGS = config["tags"]  # list of tags to search for in calendar event
SCOPES = [config["scopes"]]  # api adress to access calendar data

def excel_homeoffice_setter(ws2,day):
    d = 0
    for day, tags in vor_pro.items():
        if tags["vor"] and not tags["pro"]:
            d += 1
            excel_setter_homeoffice_p(ws2,datum=day, d=d)

def vor_counter(searchtag, not_searchtag):
    d = 0
    for day,tags in vor_pro.items():
       # print(tags)
        if tags[searchtag] and not tags[not_searchtag]:
            d += 1
    return d
def pro_counter(searchtag):
    d = 0
    for day,tags in vor_pro.items():
       # print(tags)
        if tags[searchtag]:
            d += 1
    return d

            #print(day, "nur #vor")edentials möglichst früh initialisieren, direkt nach SCOPES
# laden der credentials aus der json datei und Berechtigungen für API
credentials = service_account.Credentials.from_service_account_file(
    path_resolver.service_account_file, scopes=SCOPES
)
# erstellt einen Dienst, um mit der Google Calendar API zu kommunizieren
service = build("calendar", "v3", credentials=credentials)

# excel laden und weitere Variablen
wb = load_workbook(path_resolver.excel_load_path)
wb2 = load_workbook(path_resolver.excel_load_path_ho)  # excel laden
ws = wb.active  # excel aktiv schalten
ws2 = wb2.active  # excel2 aktiv schalten
First_day = ""
Last_day = ""
start_date = ""
hours_prepared=0
hours_teaching=0

# Aktuellen Monat berechnen
#now = datetime(2026,4,1) #bt jahr monat und Tag an

BERLIN = ZoneInfo("Europe/Berlin")
now = datetime.now(BERLIN)

first_local = datetime(now.year, now.month, 1, 0, 0, 0, tzinfo=BERLIN)
last_local = datetime(
    now.year,
    now.month,
    calendar.monthrange(now.year, now.month)[1],
    23, 59, 59,
    tzinfo=BERLIN
)

first_utc = first_local.astimezone(ZoneInfo("UTC"))
last_utc = last_local.astimezone(ZoneInfo("UTC"))

time_min = first_utc.isoformat().replace("+00:00", "Z")
time_max = last_utc.isoformat().replace("+00:00", "Z")

events_result = service.events().list(
    calendarId=CALENDAR_ID,
    timeMin=time_min,
    timeMax=time_max,
    singleEvents=True,
    orderBy="startTime",
    timeZone = "Europe/Berlin"
).execute()

# gibt die Termine zurück, die im aktuellen Monat liegen als Liste von Ereignissen zurück. Jedes Ereignis enthält Informationen wie Start- und Endzeit, Titel, Beschreibung usw.
# wenn in der Liste der Ereignisse keine Termine gefunden werden, wird eine Nachricht ausgegeben, dass keine Termine in diesem Monat vorhanden sind. Andernfalls wird für jedes Ereignis in der Liste eine Schleife durchlaufen, um die relevanten Informationen zu extrahieren und in die Excel-Datei einzutragen.

events = events_result.get("items", [])
if not events:
    print("Keine Termine in diesem Monat.")
else:
    i = 0
  
    for event in events:
           
        start = event["start"].get("dateTime", event["start"].get("date"))
        end = event["end"].get("dateTime", event["end"].get("date"))

        summary = event.get("summary", "(kein Titel)",)
        daylie_date = start[:10]
        
        description = event.get("description", "")
        start_dt = parser.parse(start) 
        end_dt = parser.parse(end) # Datum und Uhrzeit getrennt formatieren
        start_date = start_dt.strftime("%d.%m.%Y") 
            # wenn der startwert für die Excel 32 ist dann setze das Firstdate (der Tag an dem das erste mal ein Termin oder die Vorbereitung stattfindet)
        if i == 0:
            First_day = start_date
        
        # wenn einer der Tags in den Überschriften ist, dann führe das untere aus
        if any(tag in summary for tag in TAGS):
                            
            if daylie_date not in vor_pro:
                vor_pro[daylie_date] = {"vor": False, "pro": False} 
            if ("#vor" in summary or "#Vor" in summary):
                vor_pro[daylie_date]["vor"] = True
                       
            if ("#pro" in summary or "#Pro" in summary):
                vor_pro[daylie_date]["pro"] = True

            
            # passe immer das Lastday an das start_date an, somit wird der letzte Tag durchgehen ermittelt     
            Last_day = start_date 
            start_time = start_dt.strftime("%H:%M") 
            end_date = end_dt.strftime("%d.%m.%Y") 
            end_time = end_dt.strftime("%H:%M")

            #berechnet die Stundendiferenz
            dif =  end_dt - start_dt 
            decimal_hours = dif.total_seconds() / 3600
            total_minutes = int(dif.total_seconds() // 60) 
            hours = total_minutes // 60
        
            """---------------------nochmal überarbeiten"""
            if "#vor" in summary or "#Vor" in summary: # wenn #vor in der Überschrift ist, dann addiere die Stunden zu den Vorbereitungstunden
                hours_prepared += 1
            if "#pro" in summary or "#Pro" in summary: # wenn #pro in der Überschrift ist, dann addiere die Stunden zu den Unterrichtsstunden
                hours_teaching += hours
            """--------------------------------------------"""
            #print(hours_prepared, hours_teaching)
            minutes = total_minutes % 60
            month = datetime.now().strftime("%B")
            #month = datetime(2026,3,1).strftime("%B")
            #print(f"{decimal_hours} Stunden")
            
          
            excel_setter(i+32,ws, datum=start_date, decimal_hours=decimal_hours, description=description,First_day=First_day, Last_day=Last_day) 
            wb.save(path_resolver.output_dir / f"Muster_Honorarrechnung-Lehrkräfte_{month}.xlsx")
            i += 1

"""muss nochmal angepasst werden-------------------------------"""
Last_day = datetime.strptime(Last_day, "%d.%m.%Y")
Last_day_adjusted = Last_day.strftime("%Y-%m-%d")
First_day = datetime.strptime(First_day, "%d.%m.%Y")
First_day_adjusted = First_day.strftime("%Y-%m-%d") 
now = datetime.now().strftime("%Y-%m-%d")
""""------------------------------------------------------------"""


"""wenn ordner existiert und datei existiert, lese die datei aus. wenn nichts in der datei steht erstelle eine Rechnung
wenn deine nummer existiert. lösche die rechnung und erstelle dann eine neue rechnung
"""
def delete_invoice_if_exists():
    if (path_resolver.output_dir / "last_invoice_ID.txt").exists():
        with open(path_resolver.output_dir / "last_invoice_ID.txt", "r") as f:
            content = f.read()
        if content:
            FastBill_invoice_creator_instance.delete_invoice(content)
    """logging hier einfügen"""
def create_new():
    responded_invoice_id = FastBill_invoice_creator_instance.create_invoice(start_date=First_day_adjusted,
                                                                    end_date=Last_day_adjusted,
                                                                    current_date=now,
                                                                    quantity_preparing=hours_prepared,
                                                                    quantity_teaching=hours_teaching)
    with open(path_resolver.output_dir / "last_invoice_ID.txt", "w") as f:
        f.write(str(responded_invoice_id))
    """logging hier einfügen"""

delete_invoice_if_exists()
create_new()
excel_homeoffice_setter(ws2, start_date)
wb2.save(path_resolver.output_dir / f"Homeoffice_zaehler.xlsx") 