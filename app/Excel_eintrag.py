from openpyxl import load_workbook
from zoneinfo import ZoneInfo
from datetime import datetime

# Excel-Datei laden

def excel_setter(i,ws, datum, decimal_hours, description,First_day, Last_day, stundensatz=19):
    
    BERLIN = ZoneInfo("Europe/Berlin")
    now = datetime.now(BERLIN)
    
    # ab Zeile 31 befüllen
   
    ws[f"A{i}"] = datum     # Datum
    ws[f"B{i}"] = decimal_hours                # Stunden
    ws[f"C{i}"] = description        # Leistung
    ws[f"D{i}"] = stundensatz
    ws[f"E{i}"] = f"=ROUND(B{i}*$D$32,2)"
    ws["C23"] = f"{now.month-1}/{now.year}"
    ws["C24"] = f"{First_day} bis {Last_day}"
    ws["E23"] = f"{Last_day}"
    ws["A29"] = f"für die unten aufgeführten Leistungen für den Stütz- und Förderunterricht in der Ausbildungsassistenz im Projekt Assistierte Ausbildung AsA VIII berechne ich Ihnen wie folgt: "

def excel_setter_homeoffice_p(ws2,d, datum, kunde="", betrag=6):
    d = ws2.max_row + 1
    ws2[f"A{d}"] = datum
    ws2[f"B{d}"] = f"Vorbereitungstag {kunde}"
    ws2[f"C{d}"] = f"{d}"
    ws2[f"D{d}"] = f"{betrag}€"
    

    
   

   

